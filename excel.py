#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/6/18 9:34
# @Author  : Zhang Shanxiu
import openpyxl
from collections import Counter


class Xlsxs:
    def __init__(self, path):
        self.res = list()
        self.set = set()
        self.wb = openpyxl.load_workbook(path)
        self.sh = self.wb['session(排班表)']
        self.repeat = ''


    def __del__(self):
        self.wb.close()


    def get_number(self, rows, cols = [4, 12]):
        for row in range(rows[0], rows[1]):
            for col in range(cols[0], cols[1]):
                ce = self.sh.cell(row=row, column=col)
                # print(ce.value)
                if ce.value != None and ce.value != ' ':
                    self.res.append(ce.value)


    def get_numbers(self):
        self.get_number([7, 15])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        self.get_number([24, 32])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        self.get_number([35, 43])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        self.get_number([46, 54])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        self.get_number([57, 64])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        # print(len(set(self.res)))
        # print(len(self.res))
        d = dict(Counter(self.res))
        # print({key: value for key, value in d.items() if value > 1})
        repeat = [str(key) for key in d.keys() if d[key] > 1]
        self.repeat = ', '.join(repeat)
        print(len(self.res))
        # print(len(self.set))
        print(self.repeat)
        print('重复长度', len(self.repeat))
        # print(d[' '])
        self.set = set(self.res)
        print(len(self.res) == len(self.set))
        print(len(self.set))

# 把二维列表存入excel中
def writeToExcel(file_path, list):
    # total_list = [['A', 'B', 'C', 'D', 'E'], [1, 2, 4, 6, 8], [4, 6, 7, 9, 0], [2, 6, 4, 5, 8]]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '明细'
    for r in range(len(list)):
        for c in range(len(list[0])):
            ws.cell(r + 1, c + 1).value = list[r][c]
            # excel中的行和列是从1开始计数的，所以需要+1
    wb.save(file_path)  # 注意，写入后一定要保存
    print("成功写入文件: " + file_path + " !")
    return 1


def main():
    xlsx = Xlsxs('./xlsxs/RCAR 2021 no short.xlsx')
    xlsx.get_numbers()
    # l = [[105, 108, 109, 110, 116, 119, 11, 128, 129],
    #      [133, 139, 157, 161, 163, 165, 167, 169, 172],
    #      [180, 186, 196, 204, 216, 218, 224, 227, 238],
    #      [249, 255, 256, 261, 266, 276, 27, 280, 283],
    #      [284, 288, 295, 299, 2, 301, 304, 305, 306],
    #      [30, 310, 312, 314, 322, 331, 333, 39, 3],
    #      [40, 48, 53, 57, 62, 69, 73, 81, 96]]
    # writeToExcel('./xlsxs/result.xlsx', l)


if __name__ == '__main__':
    main()