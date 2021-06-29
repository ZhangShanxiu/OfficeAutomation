#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/6/20 17:16
# @Author  : Zhang Shanxiu
import openpyxl
from collections import Counter
from openpyxl.styles  import Font, Alignment


class Author:
    def __init__(self, path_all, path_rcar):
        self.wbAll = openpyxl.load_workbook(path_all)
        self.shAll = self.wbAll['registrations']
        self.wbRcar = openpyxl.load_workbook(path_rcar)
        self.shRcar = self.wbRcar['session(排班表)']
        self.res = list()
        self.dictWrite = dict()

        # 总表字典
        self.dictAll = dict()
        # print(self.shAll.max_row)
        # print(self.shAll.max_column)
        for row in range(2, self.shAll.max_row + 1):
            for col in range(3, 5):
                number = (self.shAll.cell(row=row, column=col)).value
                if number is not None and number != ' ':
                    authors = (self.shAll.cell(row=row, column=6)).value + ' ' + (self.shAll.cell(row=row, column=5)).value
                    self.dictAll[number] = authors
        print(self.dictAll)
        print(self.dictAll[272])
        print(len(self.dictAll))

    def __del__(self):
        self.wbAll.close()
        self.wbRcar.close()


    def get_number(self, rows, cols = [4, 12]):
        for row in range(rows[0], rows[1]):
            for col in range(cols[0], cols[1]):
                ce = self.shRcar.cell(row=row, column=col)
                # print(ce.value)
                if ce.value != None and ce.value != ' ':
                    self.res.append([(self.dictAll[ce.value]).title(), self.shRcar.cell(row=row, column=1).value])
                    
    def save2excel(self, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '明细'
        self.res.sort()
        for i in range(len(self.res)):
            if self.res[i][0][0] not in self.dictWrite:
                self.dictWrite[self.res[i][0][0]] = list()
            self.dictWrite[self.res[i][0][0]].append(self.res[i])
        row, col = 1, 1
        for key, value in self.dictWrite.items():
            ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + 1)
            ws.cell(row, col).value = '--' + value[0][0][0] + '--'
            ws.cell(row, col).font = Font(name='黑体',size=18,bold=True,italic=True,color='000000')
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
            row += 2
            for r in range(len(value)):
                for c in range(len(value[0])):
                    # excel中的行和列是从1开始计数的，所以需要+1
                    ws.cell(row, c + 1).value = value[r][c]
                row += 1
        wb.save(file_path)  # 注意，写入后一定要保存
        print("成功写入文件: " + file_path + " !")
        return 1


def main():
    author = Author('./xlsxs/registrations.xlsx', './xlsxs/RCAR 2021 with short.xlsx')
    author.get_number([7, 15])
    author.get_number([24, 32])
    author.get_number([35, 43])
    author.get_number([46, 54])
    author.get_number([57, 64])
    # print(author.res)
    author.save2excel('./xlsxs/authors3.xlsx')


if __name__ == '__main__':
    main()