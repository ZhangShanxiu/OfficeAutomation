#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/6/18 22:04
# @Author  : Zhang Shanxiu
import os
import shutil
import openpyxl
from collections import Counter


class Classification:
    def __init__(self, input_path, output_path):
        self.res = set()
        self.wb = openpyxl.load_workbook(input_path)
        self.sh = self.wb['session(排班表)']
        self.input_path = './jpgs'
        self.output_path = output_path

    def __del__(self):
        self.wb.close()

    def deal_block(self, rows, cols = [4, 12]):
        for row in range(rows[0], rows[1]):
            if self.sh.cell(row, 3).value != None and self.sh.cell(row, 1).value != None:
                dir_name = self.sh.cell(row, 1).value + '_' + self.sh.cell(row, 3).value
                # dir_path = os.path.abspath(self.output_path) + '/' + dir_name
                dir_path = os.path.join(os.path.abspath(self.output_path), dir_name)
                if not os.path.isdir(dir_path):
                    os.makedirs(dir_path)
                for col in range(cols[0], cols[1]):
                    ce = self.sh.cell(row=row, column=col)
                    # print(ce.value)
                    if ce.value != None:
                        old_name = os.path.join(self.input_path, str(ce.value) + '.JPG')
                        # print(old_name)
                        new_name = os.path.join(dir_path, str(ce.value) + '.JPG')
                        # print(new_name)
                        # self.res.append(ce.value)
                        shutil.move(old_name, new_name)

    def deal_blocks(self):
        self.deal_block([7, 15])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        self.deal_block([24, 32])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        self.deal_block([35, 51])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        self.deal_block([54, 62])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        self.deal_block([66, 73])
        # print(len(self.res))
        # print('*****************************************************')
        # print(self.res)

        # print(len(set(self.res)))
        # print(len(self.res))
        # d = dict(Counter(self.res))
        # print({key: value for key, value in d.items() if value > 1})
        # repeat = [str(key) for key in d.keys() if d[key] > 1]
        # self.repeat = ', '.join(repeat)
        # print(self.repeat)
        # self.set = set(self.res)
        # print(len(self.set))


def main():
    classes = Classification('./xlsxs/RCAR 2021.xlsx', './class')
    classes.deal_blocks()


if __name__ == '__main__':
    main()